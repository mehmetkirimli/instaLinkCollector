from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
import time
import re

# =====================
# CONFIG
# =====================
EXCEL_PATH = "Go Türkiye Analiz Şablon.xlsx"
POST_LINKS_FILE = "post_links.txt"
START_ROW = 2

# =====================
# HELPERS
# =====================
def get_meta(page, prop):
    loc = page.locator(f"meta[property='{prop}']")
    return loc.first.get_attribute("content") if loc.count() else ""

def parse_desc(desc):
    likes = comments = date = ""

    m_like = re.search(r"([\d,.]+)\s+likes", desc)
    m_comment = re.search(r"([\d,.]+)\s+comments", desc)
    m_date = re.search(r"on\s+(.*?):", desc)

    if m_like:
        likes = m_like.group(1)
    if m_comment:
        comments = m_comment.group(1)
    if m_date:
        date = m_date.group(1)

    caption = desc.split(":", 1)[-1].strip() if ":" in desc else desc
    return caption, likes, comments, date

# =====================
# LOAD EXCEL
# =====================
wb = load_workbook(EXCEL_PATH)
ws = wb.active

existing_links = set()
for r in ws.iter_rows(min_row=START_ROW):
    if r[1].value:
        existing_links.add(str(r[1].value).strip())

current_row = ws.max_row + 1
print(f"📊 Excel yüklendi | Mevcut kayıt: {len(existing_links)}")

# =====================
# READ LINKS
# =====================
with open(POST_LINKS_FILE, "r", encoding="utf-8") as f:
    post_links = [l.strip() for l in f if l.strip()]

# =====================
# SCRAPE & WRITE
# =====================
with sync_playwright() as p:
    browser = p.chromium.launch(headless=False, slow_mo=40)
    context = browser.new_context(storage_state="instagram_session.json")
    page = context.new_page()

    for i, url in enumerate(post_links, start=1):
        print(f"\n➡️ [{i}] {url}")

        if url in existing_links:
            print("⏭️ Zaten var, atlandı")
            continue

        page.goto(url, timeout=60000)
        page.wait_for_load_state("domcontentloaded")
        time.sleep(2)

        desc = get_meta(page, "og:description")
        media_type = get_meta(page, "og:type")

        caption, likes, comments, date = parse_desc(desc)

        is_video = "video" in media_type.lower()

        # A - İçerik Adı / Açıklaması
        ws.cell(current_row, 1, caption)

        # B - İçerik Linki
        ws.cell(current_row, 2, url)

        # C - İçerik Tarihi
        ws.cell(current_row, 3, date)

        # D - İçerik Türü
        ws.cell(current_row, 4, "Video" if is_video else "Fotoğraf")

        if is_video:
            ws.cell(current_row, 8, likes)     # Video Beğeni
            ws.cell(current_row, 9, comments)  # Video Yorum
        else:
            ws.cell(current_row, 5, likes)     # Foto Beğeni
            ws.cell(current_row, 6, comments)  # Foto Yorum

        existing_links.add(url)
        current_row += 1
        time.sleep(2)

    browser.close()

wb.save(EXCEL_PATH)
print("\n🏁 EXCEL GÜNCELLENDİ — ADIM 5 TAMAMLANDI")
